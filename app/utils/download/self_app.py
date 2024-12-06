import argparse
import re
import sys
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class LogParser:
    def __init__(self):
        # 编译正则表达式以提高性能
        self.request_time_pattern = r"请求发起时间: (.*?)\n"
        self.response_time_pattern = r"响应时间: (.*?)\n"
        self.duration_pattern = r"请求持续时间: (.*?) 秒"
        self.response_pattern = r'{"code":(\d+).*?}(\d+)\s+([\d.]+)'

        # Excel样式定义
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.stats_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        self.header_font = Font(bold=True, size=11)
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.cell_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def parse_single_log(self, log_content: str) -> Optional[Dict]:
        """解析单条日志记录"""
        try:
            request_time = re.search(self.request_time_pattern, log_content)
            response_time = re.search(self.response_time_pattern, log_content)
            duration = re.search(self.duration_pattern, log_content)
            response = re.search(self.response_pattern, log_content)

            if not all([request_time, response_time, duration, response]):
                print(f"Warning: 某些字段未能成功匹配")
                return None

            return {
                "请求发起时间": request_time.group(1),
                "响应时间": response_time.group(1),
                "请求持续时间": float(duration.group(1)),
                "响应状态码": int(response.group(1)),
                "HTTP状态码": int(response.group(2)),
                "HTTP请求时间": float(response.group(3))
            }
        except Exception as e:
            print(f"解析错误: {str(e)}")
            return None

    def parse_log_file(self, input_file: str) -> List[Dict]:
        """解析整个日志文件"""
        results = []
        try:
            with open(input_file, 'r', encoding='utf-8') as f:
                content = f.read()
                logs = content.split("[root@")
                for log in logs:
                    if log.strip():
                        parsed = self.parse_single_log("[root@" + log)
                        if parsed:
                            results.append(parsed)
        except Exception as e:
            print(f"文件读取错误: {str(e)}")
        return results

    def save_to_excel(self, data: List[Dict], output_file: str):
        """将结果保存为Excel文件，使用公式计算统计信息"""
        if not data:
            print("没有数据要保存")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "日志分析"

            # 写入表头
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                cell.border = self.border
                ws.column_dimensions[get_column_letter(col)].width = 20

            # 写入数据
            for row, record in enumerate(data, 2):
                for col, value in enumerate(record.values(), 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(value, float):
                        cell.value = value
                        cell.number_format = '0.000000'
                    else:
                        cell.value = value
                    cell.alignment = self.cell_alignment
                    cell.border = self.border

            # 计算最后一行的数据行号
            last_row = len(data) + 1
            # 添加空行
            empty_row = last_row + 1

            # 添加统计信息（使用公式）
            stats_start_row = empty_row + 1
            stats = [
                ("统计指标", "计算值", "说明", "计算公式"),
                ("总请求数", f"=COUNT(A2:A{last_row})", "总计录入的请求数", f"COUNT函数统计A2:A{last_row}的非空单元格数"),
                ("平均请求持续时间", f'=AVERAGE(C2:C{last_row})', "所有请求的平均持续时间", f"AVERAGE函数计算C2:C{last_row}的平均值"),
                ("最长请求持续时间", f'=MAX(C2:C{last_row})', "最长的单次请求持续时间", f"MAX函数找出C2:C{last_row}中的最大值"),
                ("最短请求持续时间", f'=MIN(C2:C{last_row})', "最短的单次请求持续时间", f"MIN函数找出C2:C{last_row}中的最小值"),
                ("平均HTTP请求时间", f'=AVERAGE(F2:F{last_row})', "所有请求的平均HTTP请求时间", f"AVERAGE函数计算F2:F{last_row}的平均值"),
                ("最长HTTP请求时间", f'=MAX(F2:F{last_row})', "最长的单次HTTP请求时间", f"MAX函数找出F2:F{last_row}中的最大值"),
                ("最短HTTP请求时间", f'=MIN(F2:F{last_row})', "最短的单次HTTP请求时间", f"MIN函数找出F2:F{last_row}中的最小值"),
                ("成功请求数(HTTP 200)", f'=COUNTIF(E2:E{last_row}, 200)', "HTTP状态码为200的请求数",
                 f"COUNTIF函数统计E2:E{last_row}中等于200的单元格数"),
                ("成功请求百分比", f'=COUNTIF(E2:E{last_row}, 200)/COUNT(E2:E{last_row})*100', "成功请求占总请求的百分比",
                 "成功请求数除以总请求数乘以100")
            ]

            # 设置统计信息表头
            for col, header in enumerate(["统计指标", "计算值", "说明", "计算公式"], 1):
                cell = ws.cell(row=stats_start_row, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                cell.border = self.border
                ws.column_dimensions[get_column_letter(col)].width = 25 if col in [1, 3, 4] else 20

            # 写入统计信息
            for row_idx, stat in enumerate(stats[1:], stats_start_row + 1):
                for col_idx, value in enumerate(stat, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.border
                    cell.alignment = self.cell_alignment
                    if col_idx == 2:  # 计算值列
                        cell.number_format = '0.000000' if "HTTP" in stat[0] or "持续时间" in stat[0] else '0.00'
                    if col_idx == 1:  # 统计指标列
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.fill = self.stats_fill

            # 保存文件
            wb.save(output_file)
            print(f"数据已保存到: {output_file}")
        except Exception as e:
            print(f"保存Excel文件错误: {str(e)}")


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='日志解析工具')
    parser.add_argument('-i', '--input', required=True, help='输入日志文件路径')
    parser.add_argument('-o', '--output', required=True, help='输出文件路径')
    return parser.parse_args()


def main():
    """主函数，支持直接调用和命令行调用"""
    if len(sys.argv) > 1:
        args = parse_args()
        input_file = args.input
        output_file = args.output
    else:

        # 默认配置，用于直接调用
        # 文件路径配置
        #file_name = 'zhiguitong-2024-11-01'
        file_name = '自研新应用-2024-11-22'
        input_file = f'{file_name}.txt'
        output_file = f'{file_name}.xlsx'

        # input_file = "zhiguitong-2024-11-01-样例.log"
        # output_file = "output.xlsx"
        output_format = "xlsx"

        #input_file = "zhiguitong-2024-11-01-样例.log"
        #output_file = "output.xlsx"

    parser = LogParser()
    results = parser.parse_log_file(input_file)

    if results:
        parser.save_to_excel(results, output_file)
        print(f"\n解析完成:")
        print(f"总记录数: {len(results)}")
        print(f"输出文件: {output_file}")
    else:
        print("未找到有效的日志记录")


if __name__ == "__main__":
    main()